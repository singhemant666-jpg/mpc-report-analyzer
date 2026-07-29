import subprocess
import sys
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])

from openpyxl import load_workbook

wb = load_workbook('c:/Users/DELL/Documents/report of mpc/DAILY_COLLECTION_REPORT_2026-07-28T08-13-25-071Z.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Rows: {ws.max_row}, Columns: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True)):
        print(f'Row {i+1}: {list(row)}')
