import openpyxl

wb = openpyxl.load_workbook('c:/Users/DELL/Documents/report of mpc/DAILY_COLLECTION_REPORT_2026-07-30T06-20-15-267Z.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb['CART_PACKAGE_FINANCIAL_REPORT']

print("\n--- BHAVNA THADESHWAR RAW ROWS ---")
for row in ws.iter_rows(min_row=2, values_only=True):
    name = str(row[4] or '')
    if 'bhavna' in name.lower() or 'sumit' in name.lower():
        print(f"Date: {row[0]}, Name: {row[4]}, Item: {row[14]}, Inv: {row[19]}, Cost: {row[6]}, Paid: {row[10]}")
