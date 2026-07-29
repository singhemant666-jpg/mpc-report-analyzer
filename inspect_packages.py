import openpyxl

wb = load_workbook('c:/Users/DELL/Documents/report of mpc/DAILY_COLLECTION_REPORT_2026-07-28T08-13-25-071Z.xlsx', data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

pkg_idx = None
for i, h in enumerate(headers):
    if h and 'package' in str(h).lower():
        pkg_idx = i
        break

if pkg_idx is None:
    for i, h in enumerate(headers):
        if h and ('service' in str(h).lower() or 'item' in str(h).lower() or 'name' in str(h).lower()):
            pkg_idx = i
            break

packages = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if pkg_idx is not None and pkg_idx < len(row) and row[pkg_idx]:
        packages.add(str(row[pkg_idx]).strip())

print(f"\nTotal unique package names: {len(packages)}")
print("All package names:")
for p in sorted(packages):
    print(" -", p)
