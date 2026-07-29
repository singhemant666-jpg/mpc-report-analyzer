import openpyxl

wb = openpyxl.load_workbook('c:/Users/DELL/Documents/report of mpc/DAILY_COLLECTION_REPORT_2026-07-28T08-13-25-071Z.xlsx', data_only=True)
ws = wb['CART_PACKAGE_FINANCIAL_REPORT']

headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
print("Headers:", headers)

print("\n--- NEELAKSHI GAHERWAR ---")
for row in ws.iter_rows(min_row=2, values_only=True):
    name = str(row[4] or '') # Patient Name
    if 'neelakshi' in name.lower() or 'barik' in name.lower() or 'vaibhav' in name.lower() or 'sunny' in name.lower():
        print([str(c) if c is not None else '' for c in row])
